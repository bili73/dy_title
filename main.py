# -*- coding: utf-8 -*-
"""
main.py
================================================================================
抖音商城(livelite)「拍同款」抓取 - 运行入口（adb + RapidOCR 方案）
================================================================================
职责：解析参数 → 调用 DouyinCrawler.run() → 保存 JSON/TXT 结果。
不需要 appium server，纯 adb + Python + OCR。

运行：
    uv run python main.py
    uv run python main.py --image "D:\\xxx\\鞋.jpg"   # 临时覆盖待搜索图片
"""

import os
import sys
import json
import argparse
import logging

import config
from douyin_crawler import DouyinCrawler


def parse_args():
    """解析命令行参数：支持临时覆盖待搜索的本地图片路径。"""
    parser = argparse.ArgumentParser(
        description="抖音商城「拍同款」抓取 (adb + RapidOCR)"
    )
    parser.add_argument(
        "--image",
        default=None,
        help="待搜索的本地图片绝对路径（覆盖 config.CRAWL_CONFIG.search_image_path）",
    )
    parser.add_argument(
        "--detail",
        action="store_true",
        help="进商品详情页抓取完整标题 + 完整参数(默认只抓列表标题/价格/店铺)",
    )
    parser.add_argument(
        "--max-goods",
        type=int,
        default=5,
        help="--detail 模式下抓取的商品数量(列表第一屏前 N 个，默认 5)",
    )
    parser.add_argument(
        "--append",
        action="store_true",
        help="追加模式：把本次结果追加到现有 JSON/Excel 后(默认覆盖)",
    )
    parser.add_argument(
        "--params-keywords",
        default="",
        help="--detail 模式参数容器关键词，逗号分隔(如 '维修方式,上市时间')。"
             "OCR 命中即点参数摘要行进完整参数页(通吃齿轮/列表/表盘图标)；不填用图标模板兜底",
    )
    return parser.parse_args()


def ensure_output_dir():
    """确保结果输出目录存在。"""
    os.makedirs(config.OUTPUT_CONFIG["output_dir"], exist_ok=True)


def save_results(goods, output_dir=None, append=False):
    """将抓取结果保存为 JSON + 人类可读 txt 摘要。返回最终商品列表(供 save_excel 用)。

    output_dir: 结果目录，默认 None 读 config.OUTPUT_CONFIG（CLI 行为不变）；
                传入则用传入目录（前端覆盖）。
    append: True=追加到现有 JSON 后(不覆盖历史结果)；False=覆盖(默认)。
    """
    out_dir = output_dir or config.OUTPUT_CONFIG["output_dir"]
    os.makedirs(out_dir, exist_ok=True)
    json_path = os.path.join(out_dir, config.OUTPUT_CONFIG["output_file"])
    # 追加模式：先读现有结果，再拼上本次
    existing = []
    if append and os.path.exists(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                existing = json.load(f) or []
        except Exception:
            existing = []
    all_goods = existing + [{k: v for k, v in g.items() if k != "box"} for g in goods]
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_goods, f, ensure_ascii=False, indent=2)
    if append and existing:
        print(f"\n[保存] JSON: {json_path}（追加 {len(goods)} 件，累计 {len(all_goods)} 件）")
    else:
        print(f"\n[保存] JSON: {json_path}（共 {len(all_goods)} 件商品）")

    if config.OUTPUT_CONFIG["save_txt_summary"]:
        txt_path = os.path.join(out_dir, config.OUTPUT_CONFIG["txt_file"])
        with open(txt_path, "w", encoding="utf-8") as f:
            for idx, g in enumerate(all_goods, 1):
                f.write(f"#{idx} {g.get('title', '')}\n")
                price_line = g.get('price', '')
                if g.get('coupon_price'):
                    price_line += f"（{g['coupon_price']}）"
                f.write(f"   价格: {price_line}\n")
                f.write(f"   店铺: {g.get('shop', '')}\n")
                params = g.get('params')
                if params:
                    f.write("   参数:\n")
                    for k, v in params.items():
                        f.write(f"     {k}: {v}\n")
                f.write("\n")
        print(f"[保存] TXT: {txt_path}")
    return all_goods


def save_excel(goods, output_dir=None):
    """保存到 Excel：每商品一行，参数合并到「参数」列(换行显示 key:value)。

    output_dir: 结果目录，默认 None 读 config.OUTPUT_CONFIG；传入则用传入目录。
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font

    out_dir = output_dir or config.OUTPUT_CONFIG["output_dir"]
    os.makedirs(out_dir, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "抖音拍同款"
    ws.append(["序号", "标题", "原价", "券后价", "店铺", "参数"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for idx, g in enumerate(goods, 1):
        params = g.get("params") or {}
        param_text = "\n".join(f"{k}: {v}" for k, v in params.items())
        ws.append([idx, g.get("title", ""), g.get("price", ""),
                   g.get("coupon_price", ""), g.get("shop", ""), param_text])
    # 参数列自动换行 + 设置列宽
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col, width in zip("ABCDEF", [6, 42, 12, 14, 16, 52]):
        ws.column_dimensions[col].width = width
    path = os.path.join(out_dir, config.OUTPUT_CONFIG["excel_file"])
    try:
        wb.save(path)
    except PermissionError:
        # 文件被占用(Excel 打开着)，改用带时间戳的备选文件名，避免整个程序崩
        import time as _t
        path = os.path.join(out_dir,
                            f"douyin_results_{_t.strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(path)
        print("[警告] douyin_results.xlsx 被占用(请关闭 Excel)，已存到带时间戳的备选文件")
    print(f"[保存] Excel: {path}")


def main():
    """主流程：解析参数 → 运行抓取 → 保存结果。"""
    args = parse_args()
    if args.image:
        config.CRAWL_CONFIG["search_image_path"] = os.path.abspath(args.image)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    print("=" * 70)
    print(" 抖音商城「拍同款」抓取（adb + RapidOCR）")
    print("=" * 70)

    # 解析参数容器关键词(中英文逗号分隔 → 列表，空则 crawler 回退图标模板匹配)
    params_keywords = [k.strip() for k in args.params_keywords.replace("，", ",").split(",") if k.strip()]
    crawler = DouyinCrawler(params_keywords=params_keywords)
    try:
        goods = crawler.run_detail(args.max_goods) if args.detail else crawler.run()
    except FileNotFoundError as e:
        print(f"\n[错误] 文件不存在: {e}")
        sys.exit(2)
    except Exception as e:  # 兜底：保证异常可见而非静默吞掉
        logging.exception("抓取过程中发生未预期异常")
        sys.exit(1)

    if not goods:
        print("\n[警告] 未抓取到商品，请检查 locators 关键词或拍同款流程是否走通")
        sys.exit(0)

    all_goods = save_results(goods, append=args.append)
    save_excel(all_goods)
    print("\n抓取完成 ✅")


if __name__ == "__main__":
    main()
